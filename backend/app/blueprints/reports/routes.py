"""Routes du blueprint `reports` : indicateurs synthetiques pour le tableau de bord."""
import io
import json
import time as time_module
from datetime import datetime, date, time, timedelta

from flask import current_app, jsonify, request, stream_with_context, Response
from sqlalchemy import func
from decimal import Decimal

from app.extensions import db
from app.blueprints.reports import reports_bp
from app.models import (
    Product, Sale, SaleStatus, Stock, Customer, SaleLine, User,
    Branch, SupplierReception, SupplierReceptionLine, ReceptionStatus, PaymentType,
)
from app.services.analytics_service import compute_dashboard, compute_dashboard_realtime, top_products_for_period
from app.utils.decorators import require_permission
from app.utils.pdf import build_dashboard_report_pdf, build_credits_report_pdf, pdf_response


@reports_bp.get("/dashboard")
@require_permission("reports:read")
def dashboard_summary():
    """Indicateurs cles du jour : ventes, panier moyen, alertes de stock (RF-23)."""
    from flask_jwt_extended import get_jwt
    jwt_claims = get_jwt()
    branch_id = request.args.get("branch_id")

    if jwt_claims.get("role") == "VENDEUR":
        jwt_branch = jwt_claims.get("branch_id")
        if not jwt_branch:
            return jsonify({
                "error": "NO_BRANCH_ASSIGNED",
                "message": "Votre compte n'est rattache a aucune boutique. "
                           "Contactez votre administrateur pour qu'il vous assigne une branche.",
            }), 403
        branch_id = jwt_branch

    today_start = datetime.combine(datetime.utcnow().date(), time.min)

    sales_query = Sale.query.filter(
        Sale.status == SaleStatus.VALIDEE.value, Sale.created_at >= today_start
    )
    if branch_id:
        sales_query = sales_query.filter(Sale.branch_id == branch_id)

    sales = sales_query.all()
    ca = sum(float(s.total) for s in sales)
    nb = len(sales)
    panier = ca / nb if nb else 0

    low_stock_query = (
        Stock.query
        .join(Product, Stock.product_id == Product.id)
        .filter(Stock.quantity <= Product.min_stock_threshold)
    )
    if branch_id:
        low_stock_query = low_stock_query.filter(Stock.branch_id == branch_id)
    low_stock = low_stock_query.count()

    top_products = top_products_for_period(branch_id=branch_id, days=1, limit=5)

    return jsonify({
        "sales_today_total": str(round(ca, 2)),
        "sales_today_count": nb,
        "average_basket": str(round(panier, 2)),
        "low_stock_count": low_stock,
        "top_products_today": [
            {
                "product_id": p["product_id"],
                "name": p["product_name"],
                "sku": p["sku"],
                "quantity_sold": p["total_quantity"],
            }
            for p in top_products
        ],
    })


@reports_bp.get("/dashboard/realtime")
@require_permission("reports:read")
def dashboard_realtime():
    """Snapshot temps reel du tableau de bord (section 22.2/22.5)."""
    branch_id = request.args.get("branch_id")
    payload = compute_dashboard_realtime(branch_id=branch_id)
    return jsonify(payload)


@reports_bp.get("/dashboard/stream")
@require_permission("reports:read")
def dashboard_stream():
    """Flux temps reel (Server-Sent Events) du tableau de bord (section 22.2)."""
    branch_id = request.args.get("branch_id")

    if current_app.config.get("DISABLE_SSE"):
        payload = compute_dashboard_realtime(branch_id=branch_id)
        def single_shot():
            yield "event: sse-disabled\ndata: {}\n\n"
            yield "data: " + json.dumps(payload) + "\n\n"
        return Response(
            stream_with_context(single_shot()),
            mimetype="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    def generate():
        try:
            while True:
                payload = compute_dashboard_realtime(branch_id=branch_id)
                yield "data: " + json.dumps(payload) + "\n\n"
                time_module.sleep(30)
        except GeneratorExit:
            pass

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@reports_bp.get("/vendeur/dashboard")
@require_permission("reports:read")
def vendeur_dashboard():
    """Indicateurs de performance individuels pour le vendeur connecte."""
    from flask_jwt_extended import get_jwt_identity
    cashier_id = get_jwt_identity()
    user = User.query.get(cashier_id)
    if not user:
        return jsonify({"error": "Utilisateur introuvable"}), 404

    today = datetime.utcnow().date()
    today_start = datetime.combine(today, time.min)
    month_start = datetime.combine(today.replace(day=1), time.min)

    day_sales = Sale.query.filter(
        Sale.cashier_id == cashier_id,
        Sale.status == SaleStatus.VALIDEE.value,
        Sale.created_at >= today_start,
    ).all()

    ca_jour = sum(float(s.total) for s in day_sales)
    nb_ventes_jour = len(day_sales)
    panier_moyen = ca_jour / nb_ventes_jour if nb_ventes_jour else 0

    month_sales = Sale.query.filter(
        Sale.cashier_id == cashier_id,
        Sale.status == SaleStatus.VALIDEE.value,
        Sale.created_at >= month_start,
    ).all()

    ca_mois = sum(float(s.total) for s in month_sales)
    nb_ventes_mois = len(month_sales)

    commission_rate = float(current_app.config.get("COMMISSION_RATE", 0.02))
    objectif = float(current_app.config.get("VENDEUR_MONTHLY_TARGET", 500_000))
    commission = ca_mois * commission_rate
    progression = (ca_mois / objectif * 100) if objectif else 0

    from collections import defaultdict
    heure_ca = defaultdict(float)
    for s in day_sales:
        heure_ca[s.created_at.hour] += float(s.total)
    historique_jour = [{"heure": h, "ca": round(heure_ca[h], 2)} for h in range(24)]

    product_stats = defaultdict(lambda: {"qte": 0, "ca": 0.0, "name": ""})
    for s in month_sales:
        for line in s.lines:
            pid = line.product_id
            product_stats[pid]["qte"] += line.quantity
            product_stats[pid]["ca"] += float(line.line_total)
            product_stats[pid]["name"] = line.product.name if line.product else pid

    top_produits = sorted(product_stats.items(), key=lambda x: x[1]["ca"], reverse=True)[:5]
    top_produits_mois = [
        {
            "product_id": pid,
            "name": stats["name"],
            "qte_vendue": stats["qte"],
            "ca": round(stats["ca"], 2),
        }
        for pid, stats in top_produits
    ]

    last_sales = (
        Sale.query.filter(
            Sale.cashier_id == cashier_id,
            Sale.status == SaleStatus.VALIDEE.value,
        )
        .order_by(Sale.created_at.desc())
        .limit(10)
        .all()
    )

    dernieres_ventes = [
        {
            "id": s.id,
            "reference": s.reference,
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "customer_name": s.customer.full_name if s.customer else None,
            "payment_type": s.payment_type,
            "total": float(s.total),
            "nb_lignes": len(s.lines),
        }
        for s in last_sales
    ]

    return jsonify({
        "cashier": {
            "id": user.id,
            "full_name": user.full_name,
            "branch_name": user.branch.name if user.branch else None,
        },
        "kpis_jour": {
            "ca_jour": round(ca_jour, 2),
            "nb_ventes": nb_ventes_jour,
            "panier_moyen": round(panier_moyen, 2),
        },
        "kpis_mois": {
            "ca_mois": round(ca_mois, 2),
            "nb_ventes": nb_ventes_mois,
            "commission_estimee": round(commission, 2),
            "objectif_mensuel": round(objectif, 2),
            "progression_pct": round(progression, 2),
            "commission_rate_pct": round(commission_rate * 100, 2),
        },
        "historique_jour": historique_jour,
        "top_produits_mois": top_produits_mois,
        "dernieres_ventes": dernieres_ventes,
    })


@reports_bp.get("/export")
@require_permission("reports:read")
def export_dashboard_pdf():
    branch_id = request.args.get("branch_id")
    days = int(request.args.get("days", 30))
    data = compute_dashboard(branch_id=branch_id, days=days)
    pdf_bytes = build_dashboard_report_pdf(data)
    return pdf_response(pdf_bytes, "rapport-tableau-de-bord.pdf")


@reports_bp.get("/export/sales")
@require_permission("reports:read")
def export_sales_excel():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file

    branch_id = request.args.get("branch_id")
    days = int(request.args.get("days", 30))
    since = datetime.utcnow() - timedelta(days=days)

    q = Sale.query.filter(
        Sale.status == SaleStatus.VALIDEE.value,
        Sale.created_at >= since,
    )
    if branch_id:
        q = q.filter(Sale.branch_id == branch_id)
    sales = q.order_by(Sale.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventes"
    headers = ["Reference", "Date", "Caissier", "Client", "Paiement", "Sous-total", "Remise %", "Total"]
    ws.append(headers)
    for s in sales:
        ws.append([
            s.reference,
            s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else "",
            s.cashier.full_name if s.cashier else "",
            s.customer.full_name if s.customer else "Comptoir",
            s.payment_type,
            float(s.subtotal),
            s.discount_rate,
            float(s.total),
        ])
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].auto_size = True

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     download_name="ventes.xlsx")


@reports_bp.get("/export/stock")
@require_permission("reports:read")
def export_stock_excel():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file

    branch_id = request.args.get("branch_id")

    q = Stock.query
    if branch_id:
        q = q.filter(Stock.branch_id == branch_id)
    stocks = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"
    headers = ["Produit", "SKU", "Succursale", "Quantite", "Seuil alerte", "Valeur (FCFA)"]
    ws.append(headers)
    for s in stocks:
        p = s.product
        ws.append([
            p.name if p else "",
            p.sku if p else "",
            s.branch.name if s.branch else "",
            s.quantity,
            p.min_stock_threshold if p else 0,
            float(p.simple_price if p else 0) * s.quantity,
        ])
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].auto_size = True

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     download_name="stock.xlsx")


@reports_bp.get("/export/credits")
@require_permission("reports:read")
def export_credits_excel():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file
    from app.models.sales import CustomerPayment, CustomerPaymentStatus

    payments = CustomerPayment.query.filter(
        CustomerPayment.status == CustomerPaymentStatus.PENDING.value
    ).order_by(CustomerPayment.due_date).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credits"
    headers = ["Client", "Vente", "Montant", "Echeance", "Statut"]
    ws.append(headers)
    for p in payments:
        ws.append([
            p.customer.full_name if p.customer else "",
            p.sale.reference if p.sale else "",
            float(p.amount),
            p.due_date.isoformat() if p.due_date else "",
            p.status,
        ])
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].auto_size = True

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     download_name="credits.xlsx")


@reports_bp.get("/credits/pdf")
@require_permission("reports:read")
def export_credits_pdf():
    """Exporte le rapport PDF des encours clients (credit_balance > 0).

    build_credits_report_pdf attend des objets Customer (credit_balance,
    credit_limit, full_name, phone, customer_type) — on interroge donc
    Customer directement, et non CustomerPayment.
    """
    from decimal import Decimal as _D
    from app.models.sales import Customer, Sale, PaymentType, SaleStatus

    branch_id = request.args.get("branch_id")

    query = Customer.query.filter(Customer.credit_balance > _D("0"))

    if branch_id:
        # Pas de branch_id direct sur Customer — filtrer via les ventes a credit.
        credit_customer_ids = (
            db.session.query(Sale.customer_id)
            .filter(
                Sale.branch_id == branch_id,
                Sale.payment_type == PaymentType.CREDIT.value,
                Sale.status == SaleStatus.VALIDEE.value,
                Sale.customer_id.isnot(None),
            )
            .distinct()
            .subquery()
        )
        query = query.filter(Customer.id.in_(credit_customer_ids))

    customers = query.order_by(Customer.credit_balance.desc()).all()
    pdf_bytes = build_credits_report_pdf(customers, branch_id=branch_id)
    return pdf_response(pdf_bytes, "rapport-credits.pdf")


# ---------------------------------------------------------------------------
# Module comptabilite simplifie (RF-COMPTA-01)
# ---------------------------------------------------------------------------

@reports_bp.get("/compta/summary")
@require_permission("reports:read")
def compta_summary():
    """Bilan comptable simplifie : recettes vs depenses par succursale et periode.

    Recettes = ventes validees (CASH + CREDIT).
    Depenses = receptions fournisseurs validees (cout d'achat = qty x prix unitaire).
    Balance  = recettes - depenses (benefice brut estimatif).

    Parametres GET :
        branch_id   : filtrer par succursale (optionnel)
        date_debut  : YYYY-MM-DD (defaut : 1er jour du mois courant)
        date_fin    : YYYY-MM-DD (defaut : aujourd'hui)
    """
    branch_id = request.args.get("branch_id")
    today = datetime.utcnow().date()

    raw_debut = request.args.get("date_debut", "")
    raw_fin = request.args.get("date_fin", "")
    try:
        date_debut = date.fromisoformat(raw_debut) if raw_debut else today.replace(day=1)
        date_fin = date.fromisoformat(raw_fin) if raw_fin else today
    except ValueError:
        return jsonify({"error": "Format de date invalide. Utiliser YYYY-MM-DD."}), 400

    if date_debut > date_fin:
        return jsonify({"error": "date_debut doit etre anterieure ou egale a date_fin."}), 400

    dt_debut = datetime.combine(date_debut, time.min)
    dt_fin = datetime.combine(date_fin, time(23, 59, 59))

    branches_list = [
        {"id": b.id, "name": b.name}
        for b in Branch.query.order_by(Branch.name).all()
    ]

    sales_q = Sale.query.filter(
        Sale.status == SaleStatus.VALIDEE.value,
        Sale.created_at >= dt_debut,
        Sale.created_at <= dt_fin,
    )
    if branch_id:
        sales_q = sales_q.filter(Sale.branch_id == branch_id)
    sales = sales_q.order_by(Sale.created_at).all()

    recettes_cash = sum(float(s.total) for s in sales if s.payment_type == PaymentType.CASH.value)
    recettes_credit = sum(float(s.total) for s in sales if s.payment_type == PaymentType.CREDIT.value)
    recettes_total = recettes_cash + recettes_credit
    nb_ventes = len(sales)

    recept_q = SupplierReception.query.filter(
        SupplierReception.status == ReceptionStatus.VALIDEE.value,
        SupplierReception.received_at >= dt_debut,
        SupplierReception.received_at <= dt_fin,
    )
    if branch_id:
        recept_q = recept_q.filter(SupplierReception.branch_id == branch_id)
    receptions = recept_q.order_by(SupplierReception.received_at).all()

    depenses_total = 0.0
    for r in receptions:
        for line in r.lines:
            depenses_total += float(line.unit_purchase_price) * line.quantity
    nb_receptions = len(receptions)

    balance = recettes_total - depenses_total

    from collections import defaultdict
    day_recettes = defaultdict(float)
    day_depenses = defaultdict(float)

    for s in sales:
        d = s.created_at.date().isoformat()
        day_recettes[d] += float(s.total)

    for r in receptions:
        if r.received_at:
            d = r.received_at.date().isoformat()
            for line in r.lines:
                day_depenses[d] += float(line.unit_purchase_price) * line.quantity

    all_dates = set(day_recettes.keys()) | set(day_depenses.keys())
    evolution = sorted([
        {
            "date": d,
            "recettes": round(day_recettes[d], 2),
            "depenses": round(day_depenses[d], 2),
            "balance_jour": round(day_recettes[d] - day_depenses[d], 2),
        }
        for d in all_dates
    ], key=lambda x: x["date"])

    journal_entries = []

    for s in sales:
        customer_name = s.customer.full_name if s.customer else ""
        libelle = "Vente " + s.payment_type
        if customer_name:
            libelle = libelle + " - " + customer_name
        journal_entries.append({
            "date": s.created_at.isoformat() if s.created_at else None,
            "type": "RECETTE",
            "reference": s.reference,
            "libelle": libelle,
            "montant": float(s.total),
            "branch": s.branch.name if s.branch else "",
        })

    for r in receptions:
        cout = sum(float(l.unit_purchase_price) * l.quantity for l in r.lines)
        supplier_name = r.supplier.name if r.supplier else ""
        libelle = "Reception fournisseur"
        if supplier_name:
            libelle = libelle + " - " + supplier_name
        journal_entries.append({
            "date": r.received_at.isoformat() if r.received_at else None,
            "type": "DEPENSE",
            "reference": r.reference,
            "libelle": libelle,
            "montant": round(cout, 2),
            "branch": r.branch.name if r.branch else "",
        })

    journal_entries.sort(key=lambda x: x["date"] or "")
    solde_cumul = 0.0
    for entry in journal_entries:
        if entry["type"] == "RECETTE":
            solde_cumul += entry["montant"]
        else:
            solde_cumul -= entry["montant"]
        entry["solde_cumul"] = round(solde_cumul, 2)

    return jsonify({
        "periode": {
            "debut": date_debut.isoformat(),
            "fin": date_fin.isoformat(),
        },
        "branches": branches_list,
        "recettes": {
            "total": round(recettes_total, 2),
            "cash": round(recettes_cash, 2),
            "credit": round(recettes_credit, 2),
            "nb_ventes": nb_ventes,
        },
        "depenses": {
            "total": round(depenses_total, 2),
            "nb_receptions": nb_receptions,
        },
        "balance": round(balance, 2),
        "evolution_journaliere": evolution,
        "journal": journal_entries,
    })


@reports_bp.route("/branches/compare", methods=["GET"])
@require_permission("reports:read")
def branches_compare():
    """
    Tableau de bord comparatif inter-succursales (Feature C).
    Query params: date_debut (YYYY-MM-DD), date_fin (YYYY-MM-DD)
    Returns: periode, kpis[], radar_data[], evolution[], branch_names[]
    """
    date_debut_str = request.args.get("date_debut")
    date_fin_str = request.args.get("date_fin")

    today = datetime.utcnow().date()
    date_fin = datetime.strptime(date_fin_str, "%Y-%m-%d").date() if date_fin_str else today
    date_debut = datetime.strptime(date_debut_str, "%Y-%m-%d").date() if date_debut_str else date_fin.replace(day=1)

    debut_dt = datetime.combine(date_debut, time.min)
    fin_dt = datetime.combine(date_fin, datetime.max.time())

    branches = Branch.query.filter_by(is_active=True).order_by(Branch.name).all()
    branch_names = [b.name for b in branches]

    kpis = []
    for branch in branches:
        sales = (
            Sale.query
            .filter(
                Sale.branch_id == branch.id,
                Sale.status == SaleStatus.VALIDEE.value,
                Sale.created_at >= debut_dt,
                Sale.created_at <= fin_dt,
            )
            .all()
        )
        ca = sum(float(s.total) for s in sales)
        nb_ventes = len(sales)
        panier_moyen = ca / nb_ventes if nb_ventes else 0.0

        cost = sum(
            float(sl.quantity) * float(sl.product.purchase_price or 0)
            for s in sales
            for sl in s.lines
        )
        marge_brute = ca - cost
        marge_pct = (marge_brute / ca * 100) if ca else 0.0

        client_ids = {s.customer_id for s in sales if s.customer_id}
        nb_clients_actifs = len(client_ids)

        period_days = max(1, (date_fin - date_debut).days + 1)
        top_prods = top_products_for_period(branch_id=str(branch.id), days=period_days, limit=1)
        top_product = top_prods[0]["product_name"] if top_prods else "—"

        kpis.append({
            "branch_id": str(branch.id),
            "branch_name": branch.name,
            "is_depot": branch.is_depot,
            "ca": round(ca, 2),
            "nb_ventes": nb_ventes,
            "panier_moyen": round(panier_moyen, 2),
            "marge_brute": round(marge_brute, 2),
            "marge_pct": round(marge_pct, 1),
            "nb_clients_actifs": nb_clients_actifs,
            "top_product": top_product,
        })

    metrics = ["ca", "nb_ventes", "panier_moyen", "marge_pct", "nb_clients_actifs"]
    metric_labels = {
        "ca": "CA",
        "nb_ventes": "Nb ventes",
        "panier_moyen": "Panier moyen",
        "marge_pct": "Marge %",
        "nb_clients_actifs": "Clients actifs",
    }
    radar_data = []
    for metric in metrics:
        maxima = max((k[metric] for k in kpis), default=1) or 1
        row = {"metric": metric_labels[metric]}
        for k in kpis:
            row[k["branch_name"]] = round(k[metric] / maxima * 100, 1)
        radar_data.append(row)

    evolution_map: dict = {}
    for branch in branches:
        monthly = (
            db.session.query(
                func.date_format(Sale.created_at, "%Y-%m").label("mois"),
                func.sum(Sale.total).label("ca"),
            )
            .filter(
                Sale.branch_id == branch.id,
                Sale.status == SaleStatus.VALIDEE.value,
                Sale.created_at >= debut_dt,
                Sale.created_at <= fin_dt,
            )
            .group_by("mois")
            .order_by("mois")
            .all()
        )
        for row in monthly:
            if row.mois not in evolution_map:
                evolution_map[row.mois] = {"mois": row.mois}
            evolution_map[row.mois][branch.name] = round(float(row.ca or 0), 2)

    evolution = sorted(evolution_map.values(), key=lambda x: x["mois"])

    return jsonify({
        "periode": {"debut": date_debut.isoformat(), "fin": date_fin.isoformat()},
        "kpis": kpis,
        "radar_data": radar_data,
        "evolution": evolution,
        "branch_names": branch_names,
    })
